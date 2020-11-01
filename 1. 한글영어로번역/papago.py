#-*- coding: utf-8 -*-
import os
import sys
import urllib.request
import openpyxl
import constant


def main():

    # papago request
    request = urllib.request.Request(constant.NAVER_API_URL)
    request.add_header("X-Naver-Client-Id",constant.CLIENT_ID)
    request.add_header("X-Naver-Client-Secret",constant.CLIENT_SECRET)
    
    # excel file open
    wb = openpyxl.load_workbook(constant.EXCEL_FILE)
    ws = wb.active

    col_max = 1

    while(col_max<100) :
        if(ws.cell(1,col_max).value == None):
            break

        col_max=col_max + 1

    for row in range(1, ws.max_row):
        # print(row)
        
        if(ws.cell(row,1).value == None): # A 열 비어있으면 break
            break

        for col in range(1,col_max):
            cell_value = str(ws.cell(row,col).value)

            encText = urllib.parse.quote(cell_value)
            data = "source=ko&target=en&text=" + encText # source언어 target 언어 지정
            response = urllib.request.urlopen(request, data = data.encode("utf-8"))

            if(response.getcode() == 200): # 정상 api 호출 되면
                response_body = response.read()
                result = response_body.decode('utf-8').split('translatedText')[1].split('engineType')[0].split('":"')[1].split('","')[0]
                print(result)
                ws.cell(row,col).value = result
                
            else:
                print("Error Code:" + response.getcode())
            
        print(constant.PRINT_SEPERATE_LINE)

    # excel file save & exit
    wb.save(constant.EXCEL_FILE)
    wb.close()


if __name__ == "__main__":
	main()