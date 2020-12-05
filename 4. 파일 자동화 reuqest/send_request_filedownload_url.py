#-*- coding: utf-8 -*-
import requests
import json
import datetime
import openpyxl
import os
import sys
import time

def send_request(method_name, url, param, is_urlencoded=True): # request 보냄

    now_time = time.strftime('%y-%m-%d %H:%M:%S') # 시간

    if method_name == 'GET': # GET방식인 경우
        response = requests.get(url=url, params=param)
    res = {'status_code':response.status_code, 'encoding':response.encoding, 'Content-Type': response.headers['Content-Type'], 'now_time' : now_time}
    
    if 'json' in str(response.headers['Content-Type']): # json return
        return {**res, **response.json()}
    else: # 문자열 return
        return {**res, **{'text':response.text}}

def send_message_to_slack(text): # slack에 진행상황 보고
    url = "https://hooks.slack.com/services/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"
    payload = { "text" : "@유동진 \t" + text }
    requests.post(url, json=payload)

def main():
    
    param = {}
    wb = openpyxl.load_workbook('url.xlsx')
    ws = wb.active
    col_max = 1
    col_max=col_max + 1

    # 시작시간 slack 메시지
    send_message_to_slack("=========================================")
    send_message_to_slack("start time : " + str(time.strftime('%y-%m-%d %H:%M:%S')))
    send_message_to_slack("=========================================")
    
    for row in range(1, ws.max_row):
        if(ws.cell(row,1).value == None): # A 열 비어있으면 break
            break
        for col in range(1, col_max):
            cell_value = str(ws.cell(row,col).value)
            print(cell_value)
            if(row % 1000 == 0):
                send_message_to_slack("진행률...... : " + "{0:.2f}".format((row/ws.max_row) * 1000) + "row : " + str(row))
            if(cell_value != 'None'): # 빈셀이 아니라면
                try:
                    response = send_request(method_name='GET', url=cell_value, param=param)
                    time.sleep(0.3) # sleep
                    if(response['status_code'] == 200):
                        print(str(response['status_code']) + ' : success. ' + str(response['now_time']))
                    else:
                        print(str(response['status_code']) + ' : fail')
                except Exception as e:
                    send_message_to_slack("=============== error ===================")
                    send_message_to_slack("row : " + str(row))
                    send_message_to_slack(cell_value)
                    send_message_to_slack(e)
                    send_message_to_slack("=========================================")
        print("=========================================")
    wb.close()

if __name__ == "__main__":
    main()